#!/usr/bin/env python3
"""Coleta de dados para a aba "Consulta Produto" do Agente Tributário.

Gera três arquivos, todos indexados por NCM (8 dígitos, só números):
  - data/ncm_tipi.json          -> NCM + descrição (Siscomex) + alíquota de IPI (TIPI)
  - data/cest_st_sp.json        -> CEST/segmento/MVA original (Portaria CAT 68/2019 - ICMS-ST/SP)
  - data/pis_cofins_especial.json -> NCMs com PIS/COFINS monofásico ou alíquota zero
                                      (tabelas 4.3.10/4.3.13 do SPED Contribuições), com o
                                      código da natureza da receita

Diferente do update_data.py (notícias, roda todo dia), este script busca tabelas de
referência que mudam pouco — pensado para rodar semanalmente.

Fontes oficiais:
  - NCM/descrição: API pública do Portal Único Siscomex
    (portalunico.siscomex.gov.br/classif/api/publico/nomenclatura/download/json)
  - IPI: planilha TIPI da Receita Federal (gov.br)
  - CEST/MVA: anexos da Portaria CAT 68/2019 (legislacao.fazenda.sp.gov.br)
  - PIS/COFINS monofásico/alíquota zero: tabelas do SPED Contribuições (sped.rfb.gov.br)

Os benefícios de ICMS (isenção/redução de base de cálculo dos Anexos I e II do
RICMS/SP) e a alíquota interna do ICMS NÃO são gerados por este script — ficam em
data/icms_beneficios_sample.json, uma amostra pequena curada manualmente, porque
esses anexos são texto legal (não uma tabela "NCM -> benefício") e exigem
interpretação jurídica.

Este script escreve logs de diagnóstico verbosos (contagem de itens, trechos de
resposta quando o parser não encontra nada) porque as fontes de CEST/MVA e
PIS/COFINS não puderam ser validadas ao vivo durante o desenvolvimento — a ideia é
que, se o parser não bater com o formato real na primeira execução em produção, dá
para corrigir rapidamente lendo esses logs, em vez de descobrir o formato às cegas.
"""

import json
import re
import sys
import time
import unicodedata
import urllib.error
import urllib.request
from datetime import datetime, timezone
from html import unescape
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"
NCM_TIPI_FILE = DATA_DIR / "ncm_tipi.json"
CEST_ST_FILE = DATA_DIR / "cest_st_sp.json"
PIS_COFINS_FILE = DATA_DIR / "pis_cofins_especial.json"

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
)
REQUEST_TIMEOUT = 40


def http_get(url, headers=None):
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT, **(headers or {})})
    with urllib.request.urlopen(req, timeout=REQUEST_TIMEOUT) as resp:
        return resp.read()


def strip_html(text):
    if not text:
        return ""
    text = re.sub(r"<[^>]+>", " ", text)
    text = unescape(text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def only_digits(text):
    return re.sub(r"\D", "", text or "")


def strip_accents(text):
    normalized = unicodedata.normalize("NFKD", text or "")
    return "".join(c for c in normalized if not unicodedata.combining(c))


def save_json(path, payload):
    payload = dict(payload)
    payload["last_updated"] = datetime.now(timezone.utc).isoformat(timespec="seconds")
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def load_existing_items(path):
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (FileNotFoundError, json.JSONDecodeError):
        return []
    return data.get("items") or []


def save_or_keep(path, new_items, label):
    """Salva new_items, exceto quando a coleta veio vazia e já existem dados salvos.

    Evita que uma falha transitória de rede (timeout, indisponibilidade de uma
    fonte) apague dados bons já coletados em execuções anteriores.
    """
    if new_items:
        save_json(path, {"items": new_items})
        print(f"[{label}] {len(new_items)} itens salvos")
        return
    existing = load_existing_items(path)
    if existing:
        print(
            f"[warn] [{label}] coleta retornou 0 itens (provável falha de rede); "
            f"mantendo os {len(existing)} itens já salvos em {path.name}",
            file=sys.stderr,
        )
        return
    save_json(path, {"items": []})
    print(f"[{label}] 0 itens salvos (nenhum dado existente para preservar)")


# ---------------------------------------------------------------------------
# NCM + descrição (Portal Único Siscomex)
# ---------------------------------------------------------------------------

NCM_API_URL = "https://portalunico.siscomex.gov.br/classif/api/publico/nomenclatura/download/json?perfil=PUBLICO"


def fetch_ncm_table():
    """Baixa a tabela pública de NCM (código + descrição) do Portal Único Siscomex."""
    try:
        raw = http_get(NCM_API_URL)
    except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError) as exc:
        print(f"[warn] falha ao baixar tabela NCM do Siscomex: {exc}", file=sys.stderr)
        return {}

    try:
        data = json.loads(raw)
    except json.JSONDecodeError as exc:
        print(f"[warn] resposta da tabela NCM não é JSON válido: {exc}", file=sys.stderr)
        print(f"[debug] primeiros 500 bytes: {raw[:500]!r}", file=sys.stderr)
        return {}

    rows = None
    if isinstance(data, list):
        rows = data
    elif isinstance(data, dict):
        # A resposta real observada em produção tem a lista na chave
        # "Nomenclaturas" (ao lado de metadados como "Data_Ultima_Atualizacao_NCM"
        # e "Ato"). Prioriza esse nome, mas aceita qualquer lista de dicts como
        # alternativa, caso o nome da chave mude no futuro.
        if isinstance(data.get("Nomenclaturas"), list):
            rows = data["Nomenclaturas"]
        else:
            for value in data.values():
                if isinstance(value, list) and value and isinstance(value[0], dict):
                    rows = value
                    break
    if rows is None:
        print(f"[warn] não encontrei a lista de NCMs na resposta; chaves de topo: {list(data.keys()) if isinstance(data, dict) else type(data)}", file=sys.stderr)
        return {}

    print(f"[debug] NCM Siscomex: {len(rows)} linhas brutas recebidas", file=sys.stderr)
    if rows:
        print(f"[debug] NCM Siscomex: campos do primeiro item: {list(rows[0].keys())}", file=sys.stderr)

    # Os nomes dos campos podem variar (Codigo/codigo, Descricao/descricao).
    def field(row, *names):
        for name in names:
            if name in row:
                return row[name]
        lowered = {k.lower(): v for k, v in row.items()}
        for name in names:
            if name.lower() in lowered:
                return lowered[name.lower()]
        return None

    ncm_map = {}
    for row in rows:
        codigo = only_digits(str(field(row, "codigo", "Codigo", "co_ncm") or ""))
        if len(codigo) != 8:
            continue
        descricao = strip_html(field(row, "descricao", "Descricao", "no_ncm_por") or "")
        if not descricao:
            continue
        ncm_map[codigo] = descricao

    print(f"[debug] NCM Siscomex: {len(ncm_map)} códigos de 8 dígitos com descrição", file=sys.stderr)
    return ncm_map


# ---------------------------------------------------------------------------
# IPI (TIPI - Receita Federal)
# ---------------------------------------------------------------------------

TIPI_CANDIDATE_URLS = [
    # A URL sem "/view" é a que funciona de fato (confirmado em produção); a
    # com "/view" só devolve a página HTML de pré-visualização do Plone.
    "https://www.gov.br/receitafederal/pt-br/acesso-a-informacao/legislacao/documentos-e-arquivos/tipi.xlsx",
    "https://www.gov.br/receitafederal/pt-br/acesso-a-informacao/legislacao/documentos-e-arquivos/tipi.xlsx/view",
    "https://www.gov.br/receitafederal/pt-br/acesso-a-informacao/legislacao/documentos-e-arquivos/tipi.xlsx/@@download/file",
]


def fetch_tipi_ipi_rates():
    """Baixa e faz o parse da planilha TIPI, retornando {ncm: aliquota_ipi_str}."""
    try:
        import openpyxl
    except ImportError:
        print("[warn] openpyxl não instalado; pulando alíquotas de IPI (TIPI).", file=sys.stderr)
        return {}

    raw = None
    used_url = None
    for url in TIPI_CANDIDATE_URLS:
        try:
            candidate = http_get(url)
        except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError) as exc:
            print(f"[warn] falha ao acessar {url}: {exc}", file=sys.stderr)
            continue
        if candidate[:2] == b"PK":
            raw = candidate
            used_url = url
            break
        print(f"[debug] {url} não retornou um .xlsx (assinatura: {candidate[:20]!r})", file=sys.stderr)

    if raw is None:
        print("[warn] não consegui baixar a TIPI em nenhuma das URLs candidatas.", file=sys.stderr)
        return {}

    print(f"[debug] TIPI baixada de {used_url} ({len(raw)} bytes)", file=sys.stderr)

    import io
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:  # arquivo corrompido/formato inesperado
        print(f"[warn] falha ao abrir a planilha TIPI: {exc}", file=sys.stderr)
        return {}

    sheet = wb[wb.sheetnames[0]]

    header_row_idx = None
    col_ncm = col_aliquota = None
    for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=30, values_only=True)):
        normalized = [strip_accents(strip_html(str(c)).lower()) if c else "" for c in row]
        for j, cell in enumerate(normalized):
            if "ncm" in cell and col_ncm is None:
                col_ncm = j
            if "aliquota" in cell and col_aliquota is None:
                col_aliquota = j
        if col_ncm is not None and col_aliquota is not None:
            header_row_idx = i + 1
            break

    if header_row_idx is None or col_ncm is None or col_aliquota is None:
        print(
            f"[warn] não encontrei as colunas NCM/Alíquota no cabeçalho da TIPI "
            f"(col_ncm={col_ncm}, col_aliquota={col_aliquota}). Primeiras linhas: "
            f"{[list(r) for r in sheet.iter_rows(min_row=1, max_row=5, values_only=True)]}",
            file=sys.stderr,
        )
        return {}

    print(f"[debug] TIPI: cabeçalho na linha {header_row_idx}, col_ncm={col_ncm}, col_aliquota={col_aliquota}", file=sys.stderr)

    ipi_map = {}
    for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if col_ncm >= len(row):
            continue
        codigo = only_digits(str(row[col_ncm] or ""))
        if len(codigo) != 8:
            continue
        aliquota_raw = row[col_aliquota] if col_aliquota < len(row) else None
        if isinstance(aliquota_raw, float):
            # Evita artefatos de precisão de ponto flutuante (ex.: 3.9000000000000004).
            aliquota = f"{round(aliquota_raw, 2):g}"
        else:
            aliquota = str(aliquota_raw or "").strip()
        if not aliquota:
            continue
        ipi_map[codigo] = aliquota

    print(f"[debug] TIPI: {len(ipi_map)} alíquotas de IPI extraídas", file=sys.stderr)
    return ipi_map


def build_ncm_tipi():
    ncm_map = fetch_ncm_table()
    ipi_map = fetch_tipi_ipi_rates()
    items = []
    for codigo, descricao in ncm_map.items():
        items.append({
            "ncm": codigo,
            "descricao": descricao,
            "ipi_aliquota": ipi_map.get(codigo),
        })
    return items


# ---------------------------------------------------------------------------
# CEST / MVA (Portaria CAT 68/2019 - ICMS-ST em São Paulo)
# ---------------------------------------------------------------------------

CAT68_URL = "https://legislacao.fazenda.sp.gov.br/Paginas/Portaria-CAT-68-de-2019.aspx"
LEGISLACAO_BASE = "https://legislacao.fazenda.sp.gov.br"

# O Convênio ICMS 142/2018 (CONFAZ) é a fonte nacional que define CEST + NCM
# por segmento (Anexos II a XXVI). O MVA específico de São Paulo pode ter sido
# atualizado depois por portarias estaduais — por isso o MVA aqui é só o valor
# de referência encontrado na tabela, não necessariamente o vigente (o aviso
# disso já aparece na tela de resultado).
CONVENIO_142_URL = "https://www.normaslegais.com.br/legislacao/convenio-icms-142-2018.htm"

ANEXO_LINK_RE = re.compile(r'href="([^"]*[Aa]nexo[^"]*)"[^>]*>([^<]*)</a>', re.IGNORECASE)
CEST_RE = re.compile(r"(\d{2}\.\d{3}\.\d{2})")
NCM_TEXT_RE = re.compile(r"(\d{4}\.?\d{2}\.?\d{2})")
MVA_RE = re.compile(r"(\d{1,3}(?:,\d{1,2})?)\s*%")


def extract_pdf_text(raw_bytes):
    """Extrai o texto de um PDF (bytes). Retorna None se pdfplumber não estiver
    disponível ou se o arquivo não puder ser lido (não é um PDF de texto, por
    exemplo — pode ser um PDF escaneado como imagem)."""
    try:
        import pdfplumber
    except ImportError:
        print("[warn] pdfplumber não instalado; não é possível extrair texto de PDF.", file=sys.stderr)
        return None
    try:
        import io
        with pdfplumber.open(io.BytesIO(raw_bytes)) as pdf:
            pages_text = [page.extract_text() or "" for page in pdf.pages]
        return "\n".join(pages_text)
    except Exception as exc:
        print(f"[warn] falha ao extrair texto do PDF: {exc}", file=sys.stderr)
        return None


def extract_doc_text(raw_bytes):
    """Extrai o texto de um .doc (Word 97-2003, formato binário OLE2) usando o
    utilitário externo `antiword` (instalado via apt no workflow; não existe
    biblioteca Python pura confiável para esse formato legado). Retorna None
    se o binário não estiver disponível ou a extração falhar."""
    import shutil
    import subprocess
    import tempfile

    if shutil.which("antiword") is None:
        print("[warn] antiword não instalado; não é possível extrair texto de .doc.", file=sys.stderr)
        return None
    try:
        with tempfile.NamedTemporaryFile(suffix=".doc") as tmp:
            tmp.write(raw_bytes)
            tmp.flush()
            result = subprocess.run(
                ["antiword", "-m", "UTF-8.txt", tmp.name],
                capture_output=True,
                timeout=30,
            )
        if result.returncode != 0:
            print(
                f"[warn] antiword saiu com código {result.returncode}: "
                f"{result.stderr.decode('utf-8', errors='ignore')[:500]}",
                file=sys.stderr,
            )
            return None
        return result.stdout.decode("utf-8", errors="ignore")
    except Exception as exc:
        print(f"[warn] falha ao extrair texto do .doc com antiword: {exc}", file=sys.stderr)
        return None


OLE2_SIGNATURE = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def extract_table_rows(html_text):
    """Genérico: retorna uma lista de linhas de tabela, cada uma como lista de
    texto de célula (tags já removidas). Usado tanto para CEST/MVA quanto para
    as tabelas do SPED — mais robusto do que tentar casar texto corrido, já
    que não depende de suposições sobre o layout específico de cada página."""
    rows = []
    for row_match in re.finditer(r"<tr[^>]*>(.*?)</tr>", html_text, re.IGNORECASE | re.DOTALL):
        cells = re.findall(r"<td[^>]*>(.*?)</td>", row_match.group(1), re.IGNORECASE | re.DOTALL)
        if cells:
            rows.append([strip_html(c) for c in cells])
    return rows


def fetch_cat68_annex_links():
    try:
        raw = http_get(CAT68_URL).decode("utf-8", errors="ignore")
    except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError) as exc:
        print(f"[warn] falha ao acessar a Portaria CAT 68/2019: {exc}", file=sys.stderr)
        return []

    print(f"[debug] Portaria CAT 68/2019: página principal com {len(raw)} bytes", file=sys.stderr)

    links = []
    seen = set()
    for href, label in ANEXO_LINK_RE.findall(raw):
        url = href if href.startswith("http") else f"{LEGISLACAO_BASE}/{href.lstrip('/')}"
        if url in seen:
            continue
        seen.add(url)
        links.append((url, strip_html(label)))

    print(f"[debug] Portaria CAT 68/2019: {len(links)} links de anexo encontrados", file=sys.stderr)
    if not links:
        # Diagnóstico extra: a página é SharePoint (mesmo motor do Atos.aspx),
        # então o conteúdo pode estar embutido em blocos de dados em vez de
        # <a href> simples. Procura ocorrências de "anexo"/"cest"/"mva" e
        # qualquer link para arquivo (pdf/doc/xls), para orientar o próximo ajuste.
        anexo_count = len(re.findall(r"anexo", raw, re.IGNORECASE))
        cest_count = len(re.findall(r"\bcest\b", raw, re.IGNORECASE))
        mva_count = len(re.findall(r"\bmva\b", raw, re.IGNORECASE))
        file_links = re.findall(r'href="([^"]+\.(?:pdf|docx?|xlsx?))"', raw, re.IGNORECASE)
        print(
            f"[debug] CAT 68/2019: ocorrências — 'anexo'={anexo_count}, 'cest'={cest_count}, "
            f"'mva'={mva_count}, links de arquivo (pdf/doc/xls)={len(file_links)}",
            file=sys.stderr,
        )
        if file_links:
            print(f"[debug] primeiros links de arquivo: {file_links[:10]}", file=sys.stderr)
        idx = raw.lower().find("anexo")
        if idx >= 0:
            print(f"[debug] trecho ao redor da 1ª ocorrência de 'anexo': {raw[max(0, idx - 300):idx + 300]!r}", file=sys.stderr)
        else:
            print(f"[debug] trecho da página principal (1500 chars): {raw[:1500]!r}", file=sys.stderr)
    return links


def parse_cest_rows_from_html(raw, anexo_label):
    """Best-effort: procura linhas de tabela contendo CEST + NCM (+ MVA, se houver)."""
    rows = []
    for cells_text in extract_table_rows(raw):
        if len(cells_text) < 2:
            continue
        joined = " | ".join(cells_text)

        cest_match = CEST_RE.search(joined)
        ncm_match = NCM_TEXT_RE.search(joined)
        if not cest_match or not ncm_match:
            continue

        mva_match = MVA_RE.search(joined)
        # A célula de descrição costuma ser a que tem mais letras (as outras são
        # majoritariamente números: item, CEST, NCM, MVA) — mais robusto do que
        # supor uma posição fixa de coluna, que pode variar entre anexos/fontes.
        descricao_cell = max(cells_text, key=lambda c: sum(ch.isalpha() for ch in c), default="")
        rows.append({
            "cest": cest_match.group(1),
            "ncm": only_digits(ncm_match.group(1)),
            "segmento": anexo_label,
            "descricao": descricao_cell,
            "mva_original": mva_match.group(1).replace(",", ".") if mva_match else None,
        })
    return rows


def parse_cest_rows_from_text(text, anexo_label):
    """Igual a parse_cest_rows_from_html, mas para texto corrido (sem <tr>/<td>)
    — usado para o texto extraído do PDF de anexos, que não tem estrutura de
    tabela HTML. Procura, linha a linha, CEST + NCM (+ MVA, se houver) juntos."""
    rows = []
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        cest_match = CEST_RE.search(line)
        ncm_match = NCM_TEXT_RE.search(line)
        if not cest_match or not ncm_match:
            continue
        mva_match = MVA_RE.search(line)
        rows.append({
            "cest": cest_match.group(1),
            "ncm": only_digits(ncm_match.group(1)),
            "segmento": anexo_label,
            "descricao": line[:200],
            "mva_original": mva_match.group(1).replace(",", ".") if mva_match else None,
        })
    return rows


CONVENIO_142_ANEXOS_PDF_URL = "https://www.normaslegais.com.br/legislacao/Anexos-Convenio-ICMS-142-2018.pdf"


def fetch_cest_st_sp():
    """CEST/NCM: fonte primária é o Convênio ICMS 142/2018 (nacional, define os
    segmentos/CEST/NCM sujeitos à ST em todo o país). O MVA específico de SP,
    quando presente na mesma tabela, é só um valor de referência — pode ter
    sido atualizado depois por portaria estadual (aviso já fica na tela)."""
    all_rows = []

    try:
        pdf_raw = http_get(CONVENIO_142_ANEXOS_PDF_URL)
    except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError) as exc:
        print(f"[warn] falha ao acessar o PDF de anexos do Convênio 142/2018: {exc}", file=sys.stderr)
        pdf_raw = None

    if pdf_raw is not None:
        print(f"[debug] PDF anexos Convênio 142/2018: {len(pdf_raw)} bytes, assinatura {pdf_raw[:8]!r}", file=sys.stderr)
        pdf_text = extract_pdf_text(pdf_raw)
        if pdf_text:
            rows = parse_cest_rows_from_text(pdf_text, "Convênio ICMS 142/2018 (anexos)")
            print(f"[debug] PDF anexos Convênio 142/2018: {len(pdf_text)} chars de texto extraído, {len(rows)} linhas CEST/NCM extraídas", file=sys.stderr)
            if not rows:
                print(f"[debug] trecho do texto do PDF (2000 chars): {pdf_text[:2000]!r}", file=sys.stderr)
            all_rows.extend(rows)

    try:
        raw = http_get(CONVENIO_142_URL).decode("utf-8", errors="ignore")
    except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError) as exc:
        print(f"[warn] falha ao acessar o Convênio ICMS 142/2018: {exc}", file=sys.stderr)
        raw = None

    if raw is not None:
        rows = parse_cest_rows_from_html(raw, "Convênio ICMS 142/2018")
        print(f"[debug] Convênio ICMS 142/2018: {len(raw)} bytes, {len(rows)} linhas CEST/NCM extraídas", file=sys.stderr)
        if not rows:
            table_rows = extract_table_rows(raw)
            print(f"[debug] Convênio ICMS 142/2018: {len(table_rows)} linhas de tabela encontradas (nenhuma com CEST+NCM)", file=sys.stderr)
            print(f"[debug] trecho da página (1500 chars): {raw[:1500]!r}", file=sys.stderr)
        all_rows.extend(rows)

    # Mantém a tentativa na própria página da CAT 68/2019 como complemento —
    # já confirmamos que ela normalmente não tem os anexos em <a href>
    # simples, mas isso é barato de tentar e pode mudar no futuro.
    annex_links = fetch_cat68_annex_links()
    for url, label in annex_links:
        try:
            raw = http_get(url).decode("utf-8", errors="ignore")
        except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError) as exc:
            print(f"[warn] falha ao acessar anexo {url}: {exc}", file=sys.stderr)
            continue
        rows = parse_cest_rows_from_html(raw, label)
        print(f"[debug] anexo {label!r} ({url}): {len(raw)} bytes, {len(rows)} linhas CEST/NCM extraídas", file=sys.stderr)
        all_rows.extend(rows)
        time.sleep(1)

    return all_rows


# ---------------------------------------------------------------------------
# PIS/COFINS monofásico / alíquota zero (tabelas do SPED Contribuições)
# ---------------------------------------------------------------------------

# "/arquivo/show/{id}" é uma página SPA que só carrega os dados via JS (não
# tem <tr>/<td> nem texto da tabela no HTML bruto). "/arquivo/download/{id}"
# é um padrão comum em sites gov.br para servir o arquivo de verdade (PDF ou
# xlsx) por trás dessa página — tentamos essa primeiro, com "show" como
# variante de reserva caso o padrão não valha para essas tabelas específicas.
SPED_TABELAS = [
    ("monofasico", [
        "http://sped.rfb.gov.br/arquivo/download/1638",
        "http://sped.rfb.gov.br/arquivo/show/1638",
    ]),  # Tabela 4.3.10
    ("aliquota_zero", [
        "http://sped.rfb.gov.br/arquivo/download/1643",
        "http://sped.rfb.gov.br/arquivo/show/1643",
    ]),  # Tabela 4.3.13
]

# Célula de "código da natureza da receita" isolada, e âncora no rótulo
# "natureza da receita" como estratégia alternativa para quando o código
# aparece em texto corrido em vez de uma célula própria de tabela. As tabelas
# reais do SPED (extraídas do .doc via antiword) usam a coluna "Código" com
# um número simples (ex. "100", "101"), não o formato "NN.NN" que eu tinha
# suposto inicialmente — aceita os dois formatos.
NATUREZA_CELL_RE = re.compile(r"^(?:\d{1,2}\.\d{2}|\d{1,4})$")
NATUREZA_RECEITA_RE = re.compile(
    r"natureza\s*(?:da)?\s*receita[^\d]{0,20}(\d{1,2}(?:\.\d{2})?)", re.IGNORECASE
)


def extract_ncm_natureza_from_rows(table_rows, regime):
    """Procura, célula a célula, uma célula que seja só um NCM de 8 dígitos e
    outra que seja só um código de natureza da receita (ex.: "01.01") na
    mesma linha — mais preciso do que buscar em texto corrido, porque exige
    que a célula inteira bata com o padrão (não só um trecho dela)."""
    results = []
    for cells in table_rows:
        ncm = None
        natureza = None
        for cell in cells:
            cell_stripped = cell.strip()
            digits = only_digits(cell_stripped)
            if len(digits) == 8 and NCM_TEXT_RE.fullmatch(cell_stripped):
                ncm = digits
            elif NATUREZA_CELL_RE.fullmatch(cell_stripped):
                natureza = cell_stripped
        if ncm and natureza:
            descricao = max(cells, key=lambda c: sum(ch.isalpha() for ch in c), default="")
            results.append({
                "ncm": ncm,
                "regime": regime,
                "codigo_natureza_receita": natureza,
                "descricao": descricao.strip()[:200],
            })
    return results


def parse_sped_table(raw_text, regime, rows_are_html):
    """Extrai NCM + código da natureza da receita. Tenta, em ordem, algumas
    formas de dividir o texto em "linhas de tabela" (lista de células por
    linha) até achar NCM + natureza na mesma linha: tabela HTML real; texto
    "célula | célula" (vindo de xlsx); linhas delimitadas por "|" sem espaço
    ao redor (é assim que o antiword renderiza tabelas do Word extraídas de
    .doc, ex. "|100  |Descrição...|2309.90.10   |01/2011    |"); colunas
    separadas por 2+ espaços ou tab, como variante adicional. Se nenhuma
    bater, cai para uma busca por texto corrido ancorada no rótulo "natureza
    da receita", como rede de segurança para formatos inesperados."""
    if rows_are_html:
        row_strategies = [extract_table_rows(raw_text)]
    else:
        lines = [line for line in raw_text.splitlines() if line.strip()]
        row_strategies = [
            [line.split(" | ") for line in lines],
            [[c.strip() for c in line.split("|")] for line in lines if "|" in line],
            [re.split(r"\s{2,}|\t", line.strip()) for line in lines],
        ]

    max_table_rows = 0
    for table_rows in row_strategies:
        max_table_rows = max(max_table_rows, len(table_rows))
        rows = extract_ncm_natureza_from_rows(table_rows, regime)
        if rows:
            return rows, len(table_rows)

    # Rede de segurança: busca em texto corrido.
    fallback_rows = []
    for line in raw_text.splitlines():
        ncm_match = NCM_TEXT_RE.search(line)
        codigo_match = NATUREZA_RECEITA_RE.search(line)
        if not ncm_match or not codigo_match:
            continue
        fallback_rows.append({
            "ncm": only_digits(ncm_match.group(1)),
            "regime": regime,
            "codigo_natureza_receita": codigo_match.group(1),
            "descricao": strip_html(line)[:200],
        })
    return fallback_rows, max_table_rows


def fetch_pis_cofins_especial():
    all_rows = []
    for regime, urls in SPED_TABELAS:
        rows = []
        for url in urls:
            try:
                raw = http_get(url)
            except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError) as exc:
                print(f"[warn] falha ao acessar tabela SPED ({regime}) {url}: {exc}", file=sys.stderr)
                continue

            print(f"[debug] tabela SPED {regime} ({url}): {len(raw)} bytes, assinatura {raw[:10]!r}", file=sys.stderr)

            rows_are_html = True
            raw_text = None
            if raw[:2] == b"PK":
                try:
                    import io
                    import openpyxl
                    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
                    sheet = wb[wb.sheetnames[0]]
                    text_lines = []
                    for row in sheet.iter_rows(values_only=True):
                        text_lines.append(" | ".join(str(c) for c in row if c is not None))
                    raw_text = "\n".join(text_lines)
                    rows_are_html = False
                except Exception as exc:
                    print(f"[warn] falha ao abrir tabela SPED ({regime}) como xlsx: {exc}", file=sys.stderr)
            elif raw[:5] == b"%PDF-":
                pdf_text = extract_pdf_text(raw)
                if pdf_text:
                    raw_text = pdf_text
                    rows_are_html = False
                    print(f"[debug] tabela SPED {regime} ({url}): PDF com {len(pdf_text)} chars de texto extraído", file=sys.stderr)
            elif raw[:8] == OLE2_SIGNATURE:
                # Formato binário OLE2 (Word 97-2003 .doc, ou Excel 97-2003 .xls
                # com o mesmo container). Essas tabelas do SPED são publicadas
                # como .doc, então tentamos antiword primeiro.
                doc_text = extract_doc_text(raw)
                if doc_text:
                    raw_text = doc_text
                    rows_are_html = False
                    print(f"[debug] tabela SPED {regime} ({url}): .doc (OLE2) com {len(doc_text)} chars de texto extraído via antiword", file=sys.stderr)
            else:
                raw_text = raw.decode("utf-8", errors="ignore")

            if raw_text is None:
                continue

            url_rows, table_row_count = parse_sped_table(raw_text, regime, rows_are_html)
            print(
                f"[debug] tabela SPED {regime} ({url}): {table_row_count} linhas de tabela encontradas, "
                f"{len(url_rows)} com NCM/natureza extraídos",
                file=sys.stderr,
            )
            if url_rows:
                rows = url_rows
                break  # essa URL funcionou, não precisa tentar a próxima variante

            # Se nem achamos linhas de tabela, a página pode mesmo carregar os
            # dados via JavaScript/AJAX depois do carregamento inicial — procura
            # pistas de endpoint para orientar o próximo ajuste.
            service_hints = re.findall(
                r'(?:src|href|action|url)\s*[:=]\s*"([^"]*(?:\.asmx|\.ashx|\.json|/api/)[^"]*)"',
                raw_text, re.IGNORECASE,
            )
            print(f"[debug] tabela SPED {regime} ({url}): pistas de endpoint de dados: {service_hints[:10]}", file=sys.stderr)
            print(f"[debug] trecho da tabela SPED {regime} ({url}) (1500 chars): {raw_text[:1500]!r}", file=sys.stderr)
            time.sleep(1)
        all_rows.extend(rows)
    return all_rows


def main():
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    ncm_items = build_ncm_tipi()
    save_or_keep(NCM_TIPI_FILE, ncm_items, "ncm_tipi")

    cest_rows = fetch_cest_st_sp()
    save_or_keep(CEST_ST_FILE, cest_rows, "cest_st_sp")

    pis_cofins_rows = fetch_pis_cofins_especial()
    save_or_keep(PIS_COFINS_FILE, pis_cofins_rows, "pis_cofins_especial")


if __name__ == "__main__":
    main()
